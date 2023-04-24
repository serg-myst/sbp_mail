import imaplib
import email
from email import generator
import chardet
import os
import csv
import xml.etree.ElementTree as ET
import re
from datetime import datetime
import uuid
from variables import SERVER, USER_RAIF, PASSWORD_RAIF, USER_GPB, PASSWORD_GPB, \
     FILES, LETTERS, BASE_SERVER, BASE_NAME, BASE_USER, BASE_PASSWORD
import pypyodbc
import openpyxl


def connect(USER, PASSWORD):
    try:
        mail = imaplib.IMAP4_SSL(SERVER)
        mail.login(USER, PASSWORD)
        return mail
    except imaplib.IMAP4.error as ex:
        print(ex)
        return None


def make_dirs(wd):
    dir_files = os.path.join(wd, FILES)
    if not os.path.exists(dir_files):
        os.mkdir(dir_files)
    dir_letters = os.path.join(wd, LETTERS)
    if not os.path.exists(dir_letters):
        os.mkdir(dir_letters)


def save_letter(mail_dir, filename, email_message):
    outfile_name = os.path.join(mail_dir, filename + '.eml')
    with open(outfile_name, 'w') as outfile:
        gen = generator.Generator(outfile)
        gen.flatten(email_message)


def logout(mail):
    mail.close()


def rem_file(path, filename):
    try:
        os.remove(os.path.join(path, filename))
    except FileNotFoundError as ex:
        print(ex)


def remove_mail(mail, mail_id):
    result = mail.uid('COPY', mail_id, 'INBOX/Completed')
    if result[0] == 'OK':
        mail.uid('STORE', mail_id, '+FLAGS', '(\Deleted)')
        mail.expunge()


def get_date_from_filename(filename):
    match = re.search(r'\d\d\D\d\d\D\d\d\d\d', filename)
    return match[0]


def date_form(op_date, op_form):
    return str(datetime.strptime(op_date, op_form)).replace(' ', 'T')


def sql_upd_buffer(data, dt):
    try:
        connection = pypyodbc.connect('Driver={SQL Server};'
                                      'Server=' + BASE_SERVER + ';'
                                                                'Database=' + BASE_NAME + ';'
                                                                                          'uid=' + BASE_USER + ';'
                                                                                                               'pwd=' + BASE_PASSWORD + ';')
        with connection.cursor() as cursor:
            cursor.execute("exec dbo.upd_buffer ?,?", [data, dt])
    except Exception as e:
        print(e)


def create_xml_gpb(path, filename):
    # filename = 'ibsB423.xlsx'
    workbook = openpyxl.load_workbook(os.path.join(path, filename))
    worksheet = workbook.worksheets[0]
    xml_data = ET.Element('data')
    rows = ET.SubElement(xml_data, 'rows')
    for r in range(2, worksheet.max_row):
        row = ET.SubElement(rows, 'row')
        op_merch = worksheet.cell(row=r, column=12).value  # Мерчант
        op_date = worksheet.cell(row=r, column=3).value+' '+worksheet.cell(row=r, column=4).value  # Дата операции
        op_sum = worksheet.cell(row=r, column=2).value  # Сумма операции
        op_com = worksheet.cell(row=r, column=6).value  # Сумма комиссии
        day_shift = worksheet.cell(row=r, column=8).value  # Дата смены
        op_comment = worksheet.cell(row=r, column=11).value  # Комментарий, достаем магазин

        row.set('op_merch', op_merch)
        row.set('op_date', date_form(op_date, "%d.%m.%Y %H:%M:%S"))
        row.set('shop_id', op_comment[op_comment.find('маг.') + 5:])
        row.set('op_sum', str(op_sum))
        row.set('op_com', str(op_com))
        row.set('day_shift', day_shift)
        row.set('bank_spb', 'gpb_sbp')

    mydata = ET.tostring(xml_data).decode('utf-8')
    sql_upd_buffer(mydata, 'gpb_sbp')
    rem_file(path, filename)


def create_xml_raif(path, filename):
    check_file = os.path.exists(os.path.join(path, filename))
    if check_file:
        with open(os.path.join(path, filename), 'r') as csv_file:
            reader = csv.DictReader(csv_file, delimiter=';')
            day_shift = date_form(get_date_from_filename(filename), "%d_%m_%Y")
            xml_data = ET.Element('data')
            rows = ET.SubElement(xml_data, 'rows')
            for csv_row in reader:
                row = ET.SubElement(rows, 'row')
                op_merch, op_date, op_comment, op_sum, op_com = csv_row['Мерчант'], csv_row['Дата операции МСК'], \
                                                                csv_row[
                                                                    'Комментарий'], csv_row['Сумма'], csv_row[
                                                                    'Комиссия']
                row.set('op_merch', op_merch)
                row.set('op_date', date_form(op_date, "%d.%m.%Y %H:%M:%S"))
                row.set('shop_id', op_comment[op_comment.find('СБП Маг.') + 8:])
                row.set('op_sum', op_sum)
                row.set('op_com', op_com)
                row.set('day_shift', day_shift)
                row.set('bank_spb', 'raiff_sbp')

            mydata = ET.tostring(xml_data).decode('utf-8')
            sql_upd_buffer(mydata, 'raiff_sbp')
        rem_file(path, filename)


def get_file(mail, file_type, bank):
    if mail is None:
        return

    mail.select('INBOX')
    result, data = mail.uid('search', None, "ALL")

    work_dir = os.path.abspath(os.curdir)
    make_dirs(work_dir)
    files_dir = os.path.join(work_dir, 'files')
    # mail_dir = os.path.join(work_dir, 'letters')

    if result != 'OK':
        print(f'Ошибка получения писем {result}')
        return

    if len(data) == 0:
        return

    try:
        mails_id = data[0].split()
    except:
        mails_id = data

    for id in mails_id:
        if id == b'':
            continue
        result, data = mail.uid('fetch', id, '(RFC822)')
        if result != 'OK':
            print(f'Ошибка получения письма id {id}. Результат {result}')
            continue

        email_data = data[0][1]
        result_code = chardet.detect(email_data)
        email_message = email.message_from_string(email_data.decode(result_code['encoding']))

        filename = ''
        if email_message.is_multipart():
            for part in email_message.walk():
                filename = str(part.get_filename()).replace('.'+file_type, '') + '_' + str(uuid.uuid4()) + '.'+file_type
                if part.get_content_type() == "application/octet-stream"\
                        or part.get_content_type() == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                    body = part.get_payload(decode=True)
                    with open(os.path.join(files_dir, filename), 'wb') as new_file:
                        new_file.write(body)

            if bank == 'raif':
                create_xml_raif(files_dir, filename)
            else:
                create_xml_gpb(files_dir, filename)
            # save_letter(mail_dir, filename, email_message)

        remove_mail(mail, id)

    logout(mail)


def get_emails_raif():
    mail = connect(USER_RAIF, PASSWORD_RAIF)
    get_file(mail, 'csv', 'raif')


def get_emails_gpb():
    mail = connect(USER_GPB, PASSWORD_GPB)
    get_file(mail, 'xlsx', 'gpb')


def get_emails():
    get_emails_raif()
    get_emails_gpb()


if __name__ == '__main__':
    get_emails()

    '''
    work_dir = os.path.abspath(os.curdir)
    files_dir = os.path.join(work_dir, 'files')
    create_xml_gpb(files_dir, '221216.xlsx')
    '''